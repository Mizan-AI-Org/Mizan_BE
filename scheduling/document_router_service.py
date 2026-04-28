"""
Document-to-action router.

Sibling of ``photo_router_service`` for non-image attachments. Takes a
single uploaded document (PDF / DOCX / XLSX / CSV / TXT), extracts plain
text using whatever stdlib / installed libs are available, and asks
GPT-4o to classify the document and pull out the fields a manager would
care about.

The response shape is intentionally identical to ``parse_photo`` so the
agent (and the dashboard endpoint) can branch on the same envelope.

Hard rules:
- The classifier NEVER invents a structured field. If a value isn't
  literally present in the extracted text it MUST return null.
- Confidence is conservative — anything below 0.55 means the caller
  should ask the manager instead of auto-creating a record.
- We never call the vision API on non-image files. If text extraction
  fails we return ``error`` and let the caller ask the user for the
  missing fields.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

import requests
from django.conf import settings

logger = logging.getLogger(__name__)


_VALID_CATEGORIES = {
    "invoice_or_receipt",
    "schedule",
    "id_or_certification",
    "policy_or_handbook",
    "contract",
    "report",
    "other",
}

_VALID_ACTIONS = {
    "log_invoice",
    "import_schedule",
    "upload_document",
    "ask_manager",
}

_PROMPT = """You are an AI assistant for a restaurant management system.
Below is the FULL TEXT of a business document a manager just uploaded.
Decide what kind of document it is and extract the fields a manager
would care about. ONLY use values that are literally present in the
text. If a field is not in the text, return null. Do NOT guess. Do NOT
fabricate amounts, dates, names, or invoice numbers.

Pick exactly one category from this list:
  - "invoice_or_receipt"  — supplier bill / invoice / utility bill the
                              manager owes (NOT a guest receipt the
                              customer paid)
  - "schedule"            — staff rota / weekly schedule
  - "id_or_certification" — staff ID, food handler card, license,
                              certification, training certificate
  - "policy_or_handbook"  — HR policy, employee handbook, SOP
  - "contract"            — supplier contract, lease, employment contract
  - "report"              — sales report, P&L, audit report
  - "other"               — anything that doesn't fit above

Respond with a STRICT JSON object (no markdown, no commentary) of
exactly this shape:

{
  "category": "invoice_or_receipt" | "schedule" | "id_or_certification" | "policy_or_handbook" | "contract" | "report" | "other",
  "confidence": 0.0 - 1.0,
  "summary": "one-sentence human description of the document, in English",
  "fields": {
    "vendor": string | null,
    "amount": number | null,
    "currency": string | null,
    "invoice_number": string | null,
    "due_date": "YYYY-MM-DD" | null,
    "issue_date": "YYYY-MM-DD" | null,
    "document_type": string | null,
    "person_name": string | null,
    "expiry_date": "YYYY-MM-DD" | null,
    "title": string | null
  },
  "suggested_action": "log_invoice" | "import_schedule" | "upload_document" | "ask_manager"
}

If you are not at least 55% confident, set confidence < 0.55 and
suggested_action = "ask_manager". Better to ask than to guess.
"""


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

_DOCX_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _extract_docx(blob: bytes) -> str:
    """Pure-stdlib DOCX text extraction (zip + XML)."""
    try:
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            with zf.open("word/document.xml") as fp:
                tree = ET.parse(fp)
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
        logger.warning("parse_document: docx extraction failed: %s", e)
        return ""

    parts: list[str] = []
    for paragraph in tree.iter(f"{_DOCX_NS}p"):
        chunks = [t.text or "" for t in paragraph.iter(f"{_DOCX_NS}t")]
        line = "".join(chunks).strip()
        if line:
            parts.append(line)
    return "\n".join(parts)


def _extract_xlsx(blob: bytes) -> str:
    """Excel text extraction via openpyxl (already in the venv)."""
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover
        logger.warning("parse_document: openpyxl missing")
        return ""
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as e:
        logger.warning("parse_document: xlsx open failed: %s", e)
        return ""

    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if not cells:
                continue
            parts.append(" | ".join(cells))
            row_count += 1
            if row_count >= 200:  # cap to keep the prompt small
                parts.append("...(truncated)")
                break
    try:
        wb.close()
    except Exception:
        pass
    return "\n".join(parts)


def _extract_csv(blob: bytes) -> str:
    text = _extract_text(blob) or ""
    if not text:
        return ""
    out: list[str] = []
    try:
        reader = csv.reader(io.StringIO(text))
        for i, row in enumerate(reader):
            cells = [c.strip() for c in row if c and c.strip()]
            if cells:
                out.append(" | ".join(cells))
            if i > 200:
                out.append("...(truncated)")
                break
    except Exception:
        return text[:8000]
    return "\n".join(out)


def _extract_text(blob: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return blob.decode(enc, errors="ignore")
        except Exception:
            continue
    return ""


def _extract_pdf(blob: bytes) -> str:
    """PDF extraction via pypdf if installed; otherwise empty (graceful fallback)."""
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except Exception:
            logger.info(
                "parse_document: pypdf/PyPDF2 not installed — PDF text extraction unavailable"
            )
            return ""
    try:
        reader = PdfReader(io.BytesIO(blob))
        chunks: list[str] = []
        for page in reader.pages[:30]:  # cap at 30 pages
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(c for c in chunks if c).strip()
    except Exception as e:
        logger.warning("parse_document: pdf extraction failed: %s", e)
        return ""


_DOCX_MIMES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
}
_XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
_PDF_MIMES = {"application/pdf"}
_CSV_MIMES = {"text/csv", "application/csv"}


def _detect_kind(content_type: str, name: str) -> str:
    ct = (content_type or "").lower()
    n = (name or "").lower()
    if ct in _DOCX_MIMES or n.endswith((".docx", ".doc")):
        return "docx"
    if ct in _XLSX_MIMES or n.endswith((".xlsx", ".xls")):
        return "xlsx"
    if ct in _PDF_MIMES or n.endswith(".pdf"):
        return "pdf"
    if ct in _CSV_MIMES or n.endswith(".csv"):
        return "csv"
    if ct.startswith("text/") or n.endswith((".txt", ".md", ".log")):
        return "text"
    return "unknown"


def extract_document_text(blob: bytes, content_type: str = "", name: str = "") -> tuple[str, str]:
    """Return (kind, extracted_text). kind is one of docx/xlsx/pdf/csv/text/unknown."""
    kind = _detect_kind(content_type, name)
    if kind == "docx":
        return kind, _extract_docx(blob)
    if kind == "xlsx":
        return kind, _extract_xlsx(blob)
    if kind == "pdf":
        return kind, _extract_pdf(blob)
    if kind == "csv":
        return kind, _extract_csv(blob)
    if kind == "text":
        return kind, _extract_text(blob)
    return kind, ""


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------


def _classify_text(extracted: str) -> dict[str, Any]:
    api_key = getattr(settings, "OPENAI_API_KEY", "") or ""
    if not api_key:
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": "OpenAI not configured.",
            "error": "OPENAI_API_KEY not configured",
            "suggested_action": "ask_manager",
            "fields": {},
        }

    # Cap the text we send to keep cost / latency sane.
    snippet = extracted[:15000]

    payload = {
        "model": "gpt-4o",
        "messages": [
            {"role": "system", "content": _PROMPT},
            {
                "role": "user",
                "content": (
                    "DOCUMENT TEXT BELOW (between <<<>>>). Classify and extract.\n\n"
                    f"<<<\n{snippet}\n>>>"
                ),
            },
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 800,
        "temperature": 0.1,
    }

    try:
        r = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=45,
        )
    except requests.RequestException as e:
        logger.exception("parse_document: OpenAI request failed")
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": "Document classifier failed.",
            "error": str(e),
            "suggested_action": "ask_manager",
            "fields": {},
        }

    if r.status_code != 200:
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": f"OpenAI error: {r.status_code}",
            "error": r.text[:300],
            "suggested_action": "ask_manager",
            "fields": {},
        }

    data = r.json() or {}
    text = ((data.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text)

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        logger.warning("parse_document: invalid JSON from classifier: %s", text[:300])
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": "Couldn't parse the classifier response.",
            "raw_response": text[:500],
            "suggested_action": "ask_manager",
            "fields": {},
        }

    category = str(parsed.get("category") or "other")
    if category not in _VALID_CATEGORIES:
        category = "other"
    suggested = str(parsed.get("suggested_action") or "")
    if suggested not in _VALID_ACTIONS:
        suggested = "ask_manager"

    confidence = parsed.get("confidence")
    try:
        confidence = float(confidence)
        confidence = max(0.0, min(1.0, confidence))
    except (TypeError, ValueError):
        confidence = 0.0

    return {
        "category": category,
        "confidence": confidence,
        "summary": str(parsed.get("summary") or "")[:500],
        "suggested_action": suggested,
        "fields": parsed.get("fields") or {},
    }


def parse_document(blob: bytes, content_type: str = "", name: str = "") -> dict[str, Any]:
    """Public entry point: extract + classify a non-image document."""
    kind, extracted = extract_document_text(blob, content_type=content_type, name=name)
    if kind == "unknown":
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": f"Unsupported document type ({content_type or 'unknown'}).",
            "error": "unsupported_document_type",
            "suggested_action": "ask_manager",
            "fields": {},
            "extracted_kind": kind,
        }
    if not extracted.strip():
        # File was the right type but we couldn't get any text out (e.g. PDF without
        # pypdf, scanned PDF without OCR, password-protected docx). Be honest with
        # the caller — never run the classifier on an empty string.
        return {
            "category": "other",
            "confidence": 0.0,
            "summary": "Couldn't extract any text from the document.",
            "error": "empty_extraction",
            "suggested_action": "ask_manager",
            "fields": {},
            "extracted_kind": kind,
        }

    classification = _classify_text(extracted)
    classification["extracted_kind"] = kind
    classification["extracted_chars"] = len(extracted)
    return classification
